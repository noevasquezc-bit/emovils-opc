"""
Emovils OPC — Parser de Excel de Intelcia

El call center Intelcia (servicio Punta Arrecife) envía cada tarde un Excel
con los empleados a transportar esa noche. El Excel ya viene organizado por:

  HORA → RUTA → EMPLEADOS

Ejemplo de estructura:

    Ruta 1   9:00 PM
    z01520 | Jhosias Robles | Crismeyris Sanchez Aquino | 21:00:00 | ...
    z01451 | Jhosias Robles | Eddy Jael Valdez Rudecindo | 21:00:00 | ...
    Ruta 2   9:00 PM
    z01250 | ...

Este módulo parsea el Excel y devuelve una lista estructurada de servicios
listos para insertarse en Airtable como tabla Servicios.
"""
from __future__ import annotations
import json
import logging
import re
from dataclasses import dataclass, asdict, field
from datetime import datetime, time
from pathlib import Path
from typing import Iterator, Optional

import openpyxl

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# REGEX PARA DETECTAR HEADERS DE RUTAS
# ─────────────────────────────────────────────────────────────

# Match: "Ruta 1   9:00 PM", "Ruta 2    11:00 PM", "Ruta  1   10:00 PM", etc.
RUTA_HEADER_RE = re.compile(
    r"ruta\s*(\d+)\s+(\d{1,2}):?(\d{2})?\s*([ap])m",
    re.IGNORECASE,
)

# Match Z-ID de empleado: z01520, Z01451, z00786...
Z_ID_RE = re.compile(r"^z\d{5}$", re.IGNORECASE)


# ─────────────────────────────────────────────────────────────
# ESTRUCTURAS DE DATOS
# ─────────────────────────────────────────────────────────────

@dataclass
class EmpleadoIntelcia:
    """Empleado individual con sus datos del Excel."""
    z_id: str                       # z01520
    nombre_validado: str            # nombre completo según RRHH
    direccion: str                  # dirección de RRHH
    hora_salida: str                # "21:00" formato HH:MM


@dataclass
class ServicioIntelcia:
    """Un servicio = una ruta a una hora con N empleados."""
    fecha: str                      # "2026-06-06"
    cliente: str                    # "Intelcia"
    sub_cliente: str                # "Punta Arrecife"
    ruta_id: int                    # 1, 2, 3...
    hora_servicio: str              # "21:00"
    empleados: list[EmpleadoIntelcia] = field(default_factory=list)
    cantidad_pax: int = 0
    tarifa_aplicada_rd: int = 0
    notas: str = ""

    def to_dict(self) -> dict:
        d = asdict(self)
        d["empleados_resumen"] = self.resumen_empleados()
        return d

    def resumen_empleados(self) -> str:
        """Texto compacto: '3 pax — Crismeyris S., Eddy V., ...'."""
        if not self.empleados:
            return "Sin pax"
        nombres_cortos = [
            self._nombre_corto(e.nombre_validado) for e in self.empleados[:4]
        ]
        extra = (
            f" +{len(self.empleados) - 4} más"
            if len(self.empleados) > 4 else ""
        )
        return f"{len(self.empleados)} pax — " + ", ".join(nombres_cortos) + extra

    @staticmethod
    def _nombre_corto(nombre: str) -> str:
        """'Crismeyris Sanchez Aquino' -> 'Crismeyris S.'"""
        partes = nombre.strip().split()
        if len(partes) == 1:
            return partes[0]
        return f"{partes[0]} {partes[1][0]}."


# ─────────────────────────────────────────────────────────────
# CARGA DEL TARIFARIO INTELCIA
# ─────────────────────────────────────────────────────────────

_TARIFARIO_PATH = Path(__file__).parent / "data" / "tarifario_intelcia.json"


def cargar_tarifario_intelcia() -> dict:
    """Carga el tarifario Intelcia desde el JSON."""
    with open(_TARIFARIO_PATH, encoding="utf-8") as f:
        return json.load(f)


def tarifa_para_ruta(ruta_id: int, cantidad_pax: int, tarifario: dict | None = None) -> int:
    """Devuelve la tarifa en RD$ para una ruta + cantidad de pasajeros."""
    if tarifario is None:
        tarifario = cargar_tarifario_intelcia()

    for ruta in tarifario["rutas"]:
        if ruta["ruta_id"] == ruta_id:
            if cantidad_pax <= 4:
                return ruta["tarifa_1_4_pax"]
            return ruta["tarifa_5_10_pax"]
    raise ValueError(f"Ruta {ruta_id} no existe en el tarifario Intelcia")


# ─────────────────────────────────────────────────────────────
# DETECTOR DE HEADER "Ruta X  HH:MM PM"
# ─────────────────────────────────────────────────────────────

def _parse_ruta_header(texto: str) -> Optional[tuple[int, str]]:
    """
    Si el texto es un header de ruta, devuelve (ruta_id, hora_HH:MM).
    Si no, devuelve None.
    """
    if not texto:
        return None
    m = RUTA_HEADER_RE.search(texto.strip())
    if not m:
        return None
    ruta_id = int(m.group(1))
    h = int(m.group(2))
    mm = int(m.group(3) or 0)
    ampm = m.group(4).lower()
    if ampm == "p" and h < 12:
        h += 12
    elif ampm == "a" and h == 12:
        h = 0
    return ruta_id, f"{h:02d}:{mm:02d}"


# ─────────────────────────────────────────────────────────────
# DETECTOR DE FILA DE EMPLEADO
# ─────────────────────────────────────────────────────────────

def _row_is_empleado(row: tuple) -> bool:
    """Una fila es de empleado si tiene un Z-ID en las primeras columnas."""
    if not row:
        return False
    # Z-ID puede estar en col 0 o col 1 (depende del formato del Excel)
    for idx in (0, 1):
        if idx < len(row) and row[idx]:
            if Z_ID_RE.match(str(row[idx]).strip()):
                return True
    return False


def _find_z_id_col(row: tuple) -> int:
    """Encuentra en qué columna está el Z-ID. Devuelve -1 si no hay."""
    for idx in (0, 1):
        if idx < len(row) and row[idx]:
            if Z_ID_RE.match(str(row[idx]).strip()):
                return idx
    return -1


def _normalizar_hora(valor) -> str:
    """Convierte cualquier representación de hora a HH:MM."""
    if valor is None:
        return ""
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")
    s = str(valor).strip()
    # Formatos comunes: "21:00:00", "21:00"
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


# ─────────────────────────────────────────────────────────────
# PARSER PRINCIPAL
# ─────────────────────────────────────────────────────────────

def parse_intelcia_excel(
    filepath: str | Path,
    fecha: str | None = None,
    sub_cliente: str = "Punta Arrecife",
    sheet_name: str | None = None,
) -> list[ServicioIntelcia]:
    """
    Parsea un Excel de Intelcia y devuelve una lista de Servicios estructurados.

    Args:
        filepath: Ruta al archivo .xlsx
        fecha: Fecha del servicio ("YYYY-MM-DD"). Si no se da, se infiere
               del nombre del archivo o se usa hoy.
        sub_cliente: Nombre del centro (Punta Arrecife por defecto).
        sheet_name: Nombre de la hoja a procesar. Si no, se usa la primera.

    Returns:
        Lista de ServicioIntelcia con empleados y tarifa calculada.
    """
    filepath = Path(filepath)
    if fecha is None:
        fecha = _inferir_fecha_de_archivo(filepath)

    logger.info(f"Parseando Intelcia: {filepath.name} (fecha {fecha})")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    tarifario = cargar_tarifario_intelcia()
    servicios: list[ServicioIntelcia] = []
    servicio_actual: ServicioIntelcia | None = None

    for row in ws.iter_rows(values_only=True):
        # Limpiar valores None
        celdas = [str(c).strip() if c is not None else "" for c in row]

        # Buscar header de ruta en CUALQUIER celda (puede estar en col 0 o 1)
        ruta_match = None
        for celda in celdas:
            ruta_match = _parse_ruta_header(celda)
            if ruta_match:
                break

        if ruta_match:
            # Si había un servicio en construcción, lo cerramos
            if servicio_actual and servicio_actual.empleados:
                _finalizar_servicio(servicio_actual, tarifario)
                servicios.append(servicio_actual)

            # Iniciamos nuevo servicio
            ruta_id, hora = ruta_match
            servicio_actual = ServicioIntelcia(
                fecha=fecha,
                cliente="Intelcia",
                sub_cliente=sub_cliente,
                ruta_id=ruta_id,
                hora_servicio=hora,
            )
            continue

        # Si la fila es de empleado y tenemos servicio activo, lo agregamos
        if servicio_actual and _row_is_empleado(row):
            empleado = _parse_empleado_row(row)
            if empleado:
                servicio_actual.empleados.append(empleado)

    # Cerrar el último servicio
    if servicio_actual and servicio_actual.empleados:
        _finalizar_servicio(servicio_actual, tarifario)
        servicios.append(servicio_actual)

    logger.info(f"Total servicios extraídos: {len(servicios)}")
    return servicios


def _parse_empleado_row(row: tuple) -> Optional[EmpleadoIntelcia]:
    """Convierte una fila de empleado en EmpleadoIntelcia.

    Estructura típica del Excel Intelcia (offset varía por archivo):
      col 0 o 1: Z-ID (BIOMETRIC)
      col +1: Nombre dispatch/supervisor (ej. 'Jhosias Robles')
      col +2: Nombre validado del empleado
      col +3: Hora de salida
      col +4 a +5: First name, Last name desglosados
      col +7 u +8: Dirección según RRHH
    """
    try:
        z_col = _find_z_id_col(row)
        if z_col < 0:
            return None
        z_id = str(row[z_col]).strip().lower()

        # Los demás campos están relativos al Z-ID
        offset = z_col
        nombre = (
            str(row[offset + 2]).strip()
            if len(row) > offset + 2 and row[offset + 2] else ""
        )
        hora_salida = (
            _normalizar_hora(row[offset + 3])
            if len(row) > offset + 3 else ""
        )

        # La dirección suele estar varias columnas a la derecha
        direccion = ""
        for delta in (7, 8, 6, 9):
            idx = offset + delta
            if len(row) > idx and row[idx]:
                texto = str(row[idx]).strip()
                if texto and len(texto) > 10 and texto != "#N/A":
                    direccion = texto
                    break

        if not nombre or nombre == "#N/A":
            return None

        return EmpleadoIntelcia(
            z_id=z_id,
            nombre_validado=nombre,
            direccion=direccion,
            hora_salida=hora_salida,
        )
    except Exception as e:
        logger.error(f"Error parseando empleado: {e} — row: {row}")
        return None


def _finalizar_servicio(servicio: ServicioIntelcia, tarifario: dict) -> None:
    """Calcula cantidad_pax y tarifa al cerrar un servicio."""
    servicio.cantidad_pax = len(servicio.empleados)
    try:
        servicio.tarifa_aplicada_rd = tarifa_para_ruta(
            servicio.ruta_id, servicio.cantidad_pax, tarifario
        )
    except ValueError as e:
        logger.warning(str(e))
        servicio.tarifa_aplicada_rd = 0
        servicio.notas = f"⚠️ {e}"


def _inferir_fecha_de_archivo(filepath: Path) -> str:
    """
    Intenta extraer la fecha del nombre del archivo.
    Ejemplo: "Transporte Jun-06th - 2026.xlsx" → "2026-06-06"
    Si no puede, usa la fecha de hoy.
    """
    name = filepath.stem.lower()
    meses = {
        "ene": 1, "jan": 1, "feb": 2, "mar": 3, "abr": 4, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "ago": 8, "aug": 8,
        "sep": 9, "oct": 10, "nov": 11, "dic": 12, "dec": 12,
    }
    # Buscar patrón como "Jun-06" o "Jun-06th"
    m = re.search(r"(\w{3})[-\s]?(\d{1,2})", name)
    if m:
        mes_abrev = m.group(1).lower()
        dia = int(m.group(2))
        if mes_abrev in meses:
            anio_m = re.search(r"(\d{4})", name)
            anio = int(anio_m.group(1)) if anio_m else datetime.now().year
            try:
                return datetime(anio, meses[mes_abrev], dia).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return datetime.now().strftime("%Y-%m-%d")


# ─────────────────────────────────────────────────────────────
# UTILIDADES PARA REPORTE
# ─────────────────────────────────────────────────────────────

def resumen_facturacion(servicios: list[ServicioIntelcia]) -> dict:
    """Devuelve totales del lote de servicios."""
    total_servicios = len(servicios)
    total_pax = sum(s.cantidad_pax for s in servicios)
    total_facturacion = sum(s.tarifa_aplicada_rd for s in servicios)

    por_hora: dict[str, int] = {}
    for s in servicios:
        por_hora.setdefault(s.hora_servicio, 0)
        por_hora[s.hora_servicio] += 1

    return {
        "total_servicios": total_servicios,
        "total_pasajeros": total_pax,
        "total_facturacion_rd": total_facturacion,
        "servicios_por_hora": dict(sorted(por_hora.items())),
    }


# ─────────────────────────────────────────────────────────────
# CLI PARA PRUEBAS
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    if len(sys.argv) < 2:
        # Usar el Excel real que el usuario ya nos compartió
        default = "/Users/noevasquez/Desktop/EMOVILS/Transporte Jun-06th -  2026.xlsx"
        print(f"Usando archivo por defecto: {default}")
        path = default
    else:
        path = sys.argv[1]

    servicios = parse_intelcia_excel(path)

    print("=" * 70)
    print(f"INTELCIA — {len(servicios)} SERVICIOS EXTRAÍDOS")
    print("=" * 70)

    for s in servicios:
        print(
            f"\n📍 Ruta {s.ruta_id} · {s.hora_servicio} · "
            f"{s.cantidad_pax} pax · RD${s.tarifa_aplicada_rd:,}"
        )
        for e in s.empleados[:3]:
            print(f"  • {e.z_id} — {e.nombre_validado[:35]}")
        if len(s.empleados) > 3:
            print(f"  ... +{len(s.empleados) - 3} más")

    print()
    print("=" * 70)
    resumen = resumen_facturacion(servicios)
    print(f"📊 RESUMEN")
    print(f"  Servicios totales: {resumen['total_servicios']}")
    print(f"  Pasajeros totales: {resumen['total_pasajeros']}")
    print(f"  Facturación total: RD${resumen['total_facturacion_rd']:,}")
    print(f"  Por hora: {resumen['servicios_por_hora']}")
